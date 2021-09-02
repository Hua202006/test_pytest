
import os

from openpyxl import load_workbook

from scripts.handle_path import DATA_PATH


class TestCase:
    pass


class HandleExcel:

    def __init__(self, filename, sheetname=None):
        self.filename = os.path.join(DATA_PATH, filename)
        self.sheetname = sheetname

    def read_excel(self):
        '''
        读数据
        :return:
        '''
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        testcase_list = []
        test_headers = []  # 存放表头信息
        for row in range(1, ws.max_row + 1):
            one_testcase = TestCase()  # 创建用例对象
            for column in range(1, ws.max_column + 1):
                one_cell_value = ws.cell(row, column).value
                if row == 1:
                    test_headers.append(str(one_cell_value))  # 获取表头字符串数据，方便后面调用
                else:
                    key = test_headers[column - 1]  # 获取表头字符串数据
                    if key == 'actual':
                        setattr(one_testcase, "actual_column", column)  # 动态创建列的值
                    elif key == 'result':
                        setattr(one_testcase, 'result_column', column)  # 动态创建列的值
                    setattr(one_testcase, key, one_cell_value)  # 动态创建key并赋值
            if row != 1:
                setattr(one_testcase, 'row', row)  # 动态创建每一行
                testcase_list.append(one_testcase)  # 将每个对象添加到列表

        return testcase_list

    def write_excel(self, one_testcase, actual_value, result):
        '''
        写操作
        :param one_testcase:
        :param actual_value:
        :param result:
        :return:
        '''
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        ws.cell(one_testcase.row, one_testcase.actual_column, value=actual_value)  # 获取值
        ws.cell(one_testcase.row, one_testcase.result_column, value=result)
        wb.save(self.filename)


if __name__ == '__main__':
    excel_filename = "testcase01.xlsx"
    sheet_name = "register"
    do_excel = HandleExcel(excel_filename, sheet_name)
    testcases_data = do_excel.read_excel()
    do_excel.write_excel(testcases_data[0], 10, 20)
    pass
